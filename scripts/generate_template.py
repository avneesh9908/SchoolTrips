"""
Builds the EMPTY workbook the school fills in — headers, formatting and
validation dropdowns, no invented content.

  sample-data/Trip Data.xlsx        one file, Settings + 8 source tabs
  sample-data/split/*.xlsx          the master-links-to-separate-sheets layout

Deliberately emits no rows. Shipping example trips risks the school publishing
invented content to parents, and it is the reason "dummy data" kept reappearing
in the running app.

Run:  python scripts/generate_template.py
"""

import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = pathlib.Path(__file__).resolve().parent.parent
XLSX_DIR = ROOT / "sample-data"

GRADE_CHOICES = (
    "JK,SK,Grade 1,Grade 2,Grade 3,Grade 4,Grade 5,Grade 6,"
    "Grade 7,Grade 8,Grade 9,Grade 10,Grade 11,Grade 12"
)

HEADERS = {
    "Settings": ["Key", "Link", "Notes"],
    "Students": ["StudentID", "StudentName", "Grade", "Section",
                 "ParentName", "ParentsEmailID", "FathersMobileNo", "MothersMobileNo"],
    "Trips": ["Grade", "TripTitle", "TripDates", "Status", "CoverImage", "Overview",
              "Coordinator", "CoordinatorPhone", "CoordinatorEmail", "Emergency"],
    "Itinerary": ["Grade", "Day", "Time", "Activity", "Location"],
    "Documents": ["Grade", "Label", "Url", "Category"],
    "Guidelines": ["Grade", "Type", "Text"],
    "Reminders": ["Grade", "Date", "Text"],
    "Travel": ["Grade", "Leg", "TrainNo", "Departure", "Platform", "CoachSeat", "Notes"],
    "Media": ["Grade", "Type", "Url", "Caption"],
}

SOURCES = [k for k in HEADERS if k != "Settings"]

# The Settings tab is the only one that ships rows: the keys are fixed, and a
# blank Link means "use the tab of the same name in this file".
SETTINGS_ROWS = [[k, "", f"Leave blank to use the {k} tab in this file."] for k in SOURCES]

HEADER_FILL = PatternFill("solid", fgColor="22303F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

WIDTHS = {"Text": 90, "Overview": 70, "Activity": 46, "Url": 62, "TripTitle": 44,
          "TripDates": 40, "Notes": 52, "StudentName": 22, "ParentName": 22,
          "ParentsEmailID": 30, "CoordinatorEmail": 26, "Emergency": 30, "TrainNo": 26,
          "Coordinator": 20, "CoachSeat": 34, "Label": 30, "Key": 16, "Link": 60,
          "FathersMobileNo": 18, "MothersMobileNo": 18}


def style(ws, headers, name):
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[cell.column_letter].width = WIDTHS.get(header, 16)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    if "Grade" in headers:
        col = chr(ord("A") + headers.index("Grade"))
        dv = DataValidation(type="list", formula1=f'"{GRADE_CHOICES}"', allow_blank=False)
        dv.errorTitle = "Unrecognised grade"
        dv.error = "Pick a grade from the list. A different spelling makes the row disappear from the site."
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}1000")

    if name == "Guidelines":
        dv = DataValidation(type="list", formula1='"Safety,Do,Dont,Carry"', allow_blank=False)
        dv.errorTitle = "Unrecognised type"
        dv.error = "Type must be Safety, Do, Dont or Carry."
        ws.add_data_validation(dv)
        dv.add("B2:B1000")

    if name == "Trips":
        dv = DataValidation(type="list", formula1='"Confirmed,Pending"', allow_blank=True)
        dv.errorTitle = "Unrecognised status"
        dv.error = 'Use Confirmed to show the trip as confirmed; anything else reads as "coming soon".'
        ws.add_data_validation(dv)
        col = chr(ord("A") + headers.index("Status"))
        dv.add(f"{col}2:{col}1000")


def rows_for(name):
    return SETTINGS_ROWS if name == "Settings" else []


def write_single():
    XLSX_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for name, headers in HEADERS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows_for(name):
            ws.append(row)
        style(ws, headers, name)
    path = XLSX_DIR / "Trip Data.xlsx"
    wb.save(path)
    return path


def write_split():
    out = XLSX_DIR / "split"
    out.mkdir(parents=True, exist_ok=True)
    written = []

    for name in SOURCES:
        wb = Workbook()
        ws = wb.active
        ws.title = name
        ws.append(HEADERS[name])
        style(ws, HEADERS[name], name)
        path = out / f"{name}.xlsx"
        wb.save(path)
        written.append(path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Settings"
    ws.append(HEADERS["Settings"])
    for name in SOURCES:
        ws.append([name, "", f"Paste the share link of the {name} spreadsheet here."])
    style(ws, HEADERS["Settings"], "Settings")
    path = out / "Trip Master.xlsx"
    wb.save(path)
    written.append(path)
    return written


if __name__ == "__main__":
    print(f"wrote {write_single().relative_to(ROOT)}")
    for p in write_split():
        print(f"wrote {p.relative_to(ROOT)}")

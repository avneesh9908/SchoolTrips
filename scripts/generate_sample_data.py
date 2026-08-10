"""
Generates the dummy dataset in both shapes, from one source of truth:

  sample-data/Trip Data.xlsx      upload to Drive -> becomes a Google Sheet, 8 tabs intact
  public/sample-sheets/*.csv      served by Vite, lets the real sheets adapter run locally

Run:  python scripts/generate_sample_data.py
"""

import csv
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = pathlib.Path(__file__).resolve().parent.parent
XLSX_DIR = ROOT / "sample-data"
CSV_DIR = ROOT / "public" / "sample-sheets"

GRADE_CHOICES = "JK,Grade 1,Grade 2,Grade 3,Grade 4,Grade 5,Grade 6,Grade 7,Grade 8,Grade 9,Grade 10,Grade 11,Grade 12"

# Deliberately mixed grade spellings and one blank email — the demo should prove
# the tolerances described in docs/DATA-HANDOVER.md, not hide them.
SHEETS = {
    # Optional index. Leave every Link blank and the app just reads the tabs in
    # this same file. Paste a link and that source moves to another spreadsheet,
    # with nobody touching the app's configuration.
    "Settings": (
        ["Key", "Link", "Notes"],
        [
            ["Students", "", "Leave blank to use the Students tab in this file."],
            ["Trips", "", "Leave blank to use the Trips tab in this file."],
            ["Itinerary", "", ""],
            ["Documents", "", ""],
            ["Guidelines", "", ""],
            ["Reminders", "", ""],
            ["Travel", "", ""],
            ["Media", "", ""],
        ],
    ),
    "Students": (
        ["StudentId", "StudentName", "Grade", "Section", "FatherName", "FatherEmail", "FatherPhone"],
        [
            ["S7001", "Aarav Mehta", "Grade 7", "A", "Rakesh Mehta", "rakesh.mehta@example.com", "9876543210"],
            ["S9002", "Isha Mehta", "9", "B", "Rakesh Mehta", "rakesh.mehta@example.com", "+91 98765 43210"],
            ["S7003", "Vivaan Shah", "VII", "A", "Nilesh Shah", "nilesh.shah@example.com", "9812345678"],
            ["S7004", "Diya Patel", "Grade 7", "C", "Amit Patel", "", "9900112233"],
            ["S7005", "Kabir Anand", "Grade 7", "B", "Rohit Anand", "rohit.anand@example.com", "9822334455"],
            ["S9006", "Sara Rao", "Grade 9", "A", "Sunil Rao", "sunil.rao@example.com", "9765432109"],
            ["S5007", "Myra Joshi", "Class 5", "B", "Paresh Joshi", "paresh.joshi@example.com", "9700000001"],
            ["S5008", "Aditya Nair", "Grade 5", "A", "Manoj Nair", "manoj.nair@example.com", "9700000002"],
            ["S1009", "Tara Iyer", "Grade 1", "A", "Suresh Iyer", "suresh.iyer@example.com", "9700000003"],
        ],
    ),
    "Trips": (
        ["Grade", "TripTitle", "TripDates", "Status", "CoverImage", "Overview",
         "Coordinator", "CoordinatorPhone", "CoordinatorEmail", "Emergency"],
        [
            ["Grade 7", "Jaipur · Abhaneri · Ranthambore Educational Trip",
             "Batch 1: 12–19 December 2026 · Batch 2: 13–20 December 2026", "Confirmed", "",
             "A week-long educational trip for Grade 7 covering Jaipur, Abhaneri and Ranthambore, "
             "run across two batches.\n\nPlease go through the parent and student orientation decks "
             "below before the trip. Both batches follow the same itinerary, one day apart.",
             "Ms. Anjali Desai", "+91 98200 11223", "trips.grade7@example.edu",
             "+91 98200 99887 (24×7 trip desk)"],
            ["Grade 9", "Coorg Field Study", "6–11 January 2027", "Pending", "",
             "A five-day field study on plantation ecology and water systems. Dates are provisional "
             "pending transport confirmation.",
             "Mr. Vikram Nair", "+91 98200 44556", "trips.grade9@example.edu", ""],
            ["Grade 5", "Sundarbans Nature Camp", "3–6 February 2027", "Confirmed", "",
             "A four-day introduction to mangrove ecology, with guided boat surveys and a bird count.",
             "Ms. Fatima Sheikh", "+91 98200 77665", "trips.grade5@example.edu",
             "+91 98200 99887 (24×7 trip desk)"],
        ],
    ),
    "Itinerary": (
        ["Grade", "Day", "Time", "Activity", "Location"],
        [
            ["Grade 7", "Day 1", "10:30 PM", "Depart Mumbai Central by MMCT Jaipur SF", "Mumbai Central"],
            ["Grade 7", "Day 2", "Afternoon", "Arrive Jaipur, check in, orientation walk", "Jaipur"],
            ["Grade 7", "Day 3", "Full day", "Amer Fort, Jantar Mantar and stepwell study", "Jaipur"],
            ["Grade 7", "Day 4", "Morning", "Chand Baori stepwell measurement exercise", "Abhaneri"],
            ["Grade 7", "Day 5", "Dawn", "Ranthambore safari and habitat mapping", "Ranthambore"],
            ["Grade 7", "Day 6", "Full day", "Project work and presentations at camp", "Ranthambore"],
            ["Grade 7", "Day 7", "8:20 PM", "Board JP BDTS Express for return", "Jaipur"],
            ["Grade 5", "Day 1", "7:00 AM", "Coach departs school, breakfast en route", "School"],
            ["Grade 5", "Day 2", "6:00 AM", "Sunrise boat survey and bird count", "Sundarbans"],
            ["Grade 5", "Day 3", "Full day", "Mangrove walk and sketching session", "Sundarbans"],
            ["Grade 5", "Day 4", "2:00 PM", "Return coach to school", "Sundarbans"],
        ],
    ),
    "Documents": (
        ["Grade", "Label", "Url", "Category"],
        [
            ["Grade 7", "Parent's Orientation deck",
             "https://docs.google.com/presentation/d/REPLACE_WITH_YOUR_DECK_ID/edit", "Orientation"],
            ["Grade 7", "Student's Orientation deck",
             "https://docs.google.com/presentation/d/REPLACE_WITH_YOUR_DECK_ID/edit", "Orientation"],
            ["Grade 7", "Full day-wise itinerary",
             "https://docs.google.com/document/d/REPLACE_WITH_YOUR_DOC_ID/edit", "Itinerary"],
            ["Grade 7", "Orientation poster",
             "https://drive.google.com/file/d/REPLACE_WITH_YOUR_FILE_ID/view", "Orientation"],
            ["Grade 7", "Last year's trip photos",
             "https://drive.google.com/drive/folders/REPLACE_WITH_YOUR_FOLDER_ID", "Photos"],
            ["Grade 5", "Parent briefing note",
             "https://docs.google.com/document/d/REPLACE_WITH_YOUR_DOC_ID/edit", "Orientation"],
        ],
    ),
    "Guidelines": (
        ["Grade", "Type", "Text"],
        [
            ["Grade 7", "Safety", "A ratio of 1:12 is maintained (one accompanying adult for every 12 students), including teachers, admin, support staff and Safety Monitors."],
            ["Grade 7", "Safety", "Room allocation lets teachers easily monitor students; students know which room their teacher is in."],
            ["Grade 7", "Safety", "One adult accompanies students at all times, including washroom breaks and activities."],
            ["Grade 7", "Safety", "A student orientation is held 15 days before the trip covering safety protocols."],
            ["Grade 7", "Safety", "A trained medical person and first-aid facility are available at the campsite at all times."],
            ["Grade 7", "Safety", "Nearest Police Station and Hospital details are kept handy with the Safety Monitor."],
            ["Grade 7", "Safety", "Accompanying staff receive first-aid training and carry contact details of all parents."],
            ["Grade 7", "Safety", "Staff and vendors are trained under POCSO, with signed undertakings from all vendors."],
            ["Grade 7", "Safety", "Each batch has 2 dedicated Safety Monitors focused solely on student safety."],
            ["Grade 7", "Safety", "Regular WhatsApp updates with photos and videos are shared with parents."],
            ["Grade 7", "Safety", "Headcount is taken at every stop and at regular intervals; health checks are done daily at end of day."],
            ["Grade 7", "Do", "If your child takes any medicine regularly, hand it to the Safety Monitor a day prior at the bus stop, with clear instructions."],
            ["Grade 7", "Do", "Inform the accompanying teacher if your child has any medical history needing personal attention."],
            ["Grade 7", "Dont", "Students may not carry personal medicines — staff carry a first-aid box with basics if needed."],
            ["Grade 7", "Dont", "Students may not carry mobile phones, smart watches, or other gadgets; these will be confiscated if found."],
            ["Grade 7", "Carry", "Original School ID and Aadhar card (mandatory)"],
            ["Grade 7", "Carry", "Clothes for 7 days — full pants, full-sleeve T-shirts, sweatshirts"],
            ["Grade 7", "Carry", "Sweater / fleece jacket / thermals"],
            ["Grade 7", "Carry", "Night dress"],
            ["Grade 7", "Carry", "Heavy woolens"],
            ["Grade 7", "Carry", "Shawl / light blanket"],
            ["Grade 7", "Carry", "Sport or trekking shoes (compulsory)"],
            ["Grade 7", "Carry", "Sandals / slippers"],
            ["Grade 7", "Carry", "Water bottle and a shoulder bag to carry it throughout the day"],
            ["Grade 7", "Carry", "Towel, napkin, cap or hat, bandana, cotton socks, warm gloves"],
            ["Grade 7", "Carry", "Toiletries, moisturizer, sun cream, lip balm"],
            ["Grade 7", "Carry", "Mosquito repellent cream — Odomos (compulsory)"],
            ["Grade 7", "Carry", "Dry, healthy snacks for travelling"],
            ["Grade 5", "Safety", "One adult accompanies every 8 students throughout the camp."],
            ["Grade 5", "Safety", "Life jackets are compulsory on all boat activities, without exception."],
            ["Grade 5", "Do", "Send a labelled water bottle and sun hat."],
            ["Grade 5", "Dont", "No mobile phones or smart watches."],
            ["Grade 5", "Carry", "Full-sleeve cotton clothes for 4 days"],
            ["Grade 5", "Carry", "Closed walking shoes"],
            ["Grade 5", "Carry", "Mosquito repellent (compulsory)"],
        ],
    ),
    "Reminders": (
        ["Grade", "Date", "Text"],
        [
            ["Grade 7", "1 September", "Submit signed consent form"],
            ["Grade 7", "5 September", "Submit purchase form and trip fee"],
            ["Grade 7", "10 September", "Submit medical declaration and ID copies"],
            ["Grade 5", "12 December", "Submit signed consent form"],
        ],
    ),
    "Travel": (
        ["Grade", "Leg", "TrainNo", "Departure", "Platform", "CoachSeat", "Notes"],
        [
            ["Grade 7", "Onward", "MMCT JAIPUR SF (12955)", "10:30 PM", "", "3AC — allotment shared a week prior",
             "Report at the platform 60 minutes before departure."],
            ["Grade 7", "Return", "JP BDTS EXP (12980)", "8:20 PM", "", "Overnight 3AC",
             "Arrives Surat the following morning."],
            ["Grade 5", "Onward", "Coach (chartered)", "7:00 AM", "", "Assigned seating",
             "Boarding from the school gate. Please arrive by 6:40 AM."],
        ],
    ),
    "Media": (
        ["Grade", "Type", "Url", "Caption"],
        [],
    ),
}

HEADER_FILL = PatternFill("solid", fgColor="22303F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

WIDTHS = {"Text": 90, "Overview": 70, "Activity": 46, "Url": 62, "TripTitle": 44,
          "TripDates": 40, "Notes": 42, "StudentName": 20, "FatherName": 20,
          "FatherEmail": 28, "CoordinatorEmail": 26, "Emergency": 30, "TrainNo": 26,
          "Coordinator": 20, "CoachSeat": 34, "Label": 30,
          "Key": 16, "Link": 60, "Notes": 52}


def write_xlsx():
    XLSX_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for name, (headers, rows) in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows:
            ws.append(row)

        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")
            letter = cell.column_letter
            ws.column_dimensions[letter].width = WIDTHS.get(header, 16)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

        for r in range(2, ws.max_row + 1):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

        # A dropdown on Grade is the cheapest defence against the silent
        # row-drop that an unreadable grade spelling causes.
        if "Grade" in headers:
            col = chr(ord("A") + headers.index("Grade"))
            dv = DataValidation(type="list", formula1=f'"{GRADE_CHOICES}"', allow_blank=False)
            dv.error = "Pick a grade from the list. A different spelling makes the row disappear from the site."
            dv.errorTitle = "Unrecognised grade"
            ws.add_data_validation(dv)
            dv.add(f"{col}2:{col}500")

        if name == "Guidelines":
            dv = DataValidation(type="list", formula1='"Safety,Do,Dont,Carry"', allow_blank=False)
            dv.errorTitle = "Unrecognised type"
            dv.error = "Type must be Safety, Do, Dont or Carry."
            ws.add_data_validation(dv)
            dv.add("B2:B500")

    path = XLSX_DIR / "Trip Data.xlsx"
    wb.save(path)
    return path


def write_csvs():
    CSV_DIR.mkdir(parents=True, exist_ok=True)
    written = []
    for name, (headers, rows) in SHEETS.items():
        # Settings is an index of links to other spreadsheets; it means nothing
        # in local mode, where every source is already a file on disk.
        if name == "Settings":
            continue
        path = CSV_DIR / f"{name.lower()}.csv"
        with open(path, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(headers)
            # The local fixtures get a resolvable folder id so the Drive
            # folder-expansion demo works offline; the workbook the school
            # edits keeps the obvious REPLACE_ME placeholder.
            w.writerows([[str(c).replace("REPLACE_WITH_YOUR_FOLDER_ID", "DEMO_FOLDER_ID")
                          for c in row] for row in rows])
        written.append(path)
    return written


def style_sheet(ws, headers, name):
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[cell.column_letter].width = WIDTHS.get(header, 16)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

    if "Grade" in headers:
        col = chr(ord("A") + headers.index("Grade"))
        dv = DataValidation(type="list", formula1=f'"{GRADE_CHOICES}"', allow_blank=False)
        dv.errorTitle = "Unrecognised grade"
        dv.error = "Pick a grade from the list. A different spelling makes the row disappear from the site."
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")

    if name == "Guidelines":
        dv = DataValidation(type="list", formula1='"Safety,Do,Dont,Carry"', allow_blank=False)
        dv.errorTitle = "Unrecognised type"
        dv.error = "Type must be Safety, Do, Dont or Carry."
        ws.add_data_validation(dv)
        dv.add("B2:B500")


def write_split():
    """
    The 'master sheet links to other sheets' layout: one Trip Master file whose
    Settings tab points at a separate spreadsheet per source. Only the master's
    link ever goes into the app.
    """
    out = ROOT / "sample-data" / "split"
    out.mkdir(parents=True, exist_ok=True)
    written = []

    for name, (headers, rows) in SHEETS.items():
        if name == "Settings":
            continue
        wb = Workbook()
        ws = wb.active
        ws.title = name
        ws.append(headers)
        for row in rows:
            ws.append(row)
        style_sheet(ws, headers, name)
        path = out / f"{name}.xlsx"
        wb.save(path)
        written.append(path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Settings"
    headers = ["Key", "Link", "Notes"]
    ws.append(headers)
    for key in SHEETS:
        if key == "Settings":
            continue
        ws.append([key, "", f"Paste the share link of the {key} spreadsheet here."])
    style_sheet(ws, headers, "Settings")
    path = out / "Trip Master.xlsx"
    wb.save(path)
    written.append(path)

    return written


if __name__ == "__main__":
    xlsx = write_xlsx()
    csvs = write_csvs()
    split = write_split()
    print(f"wrote {xlsx.relative_to(ROOT)}")
    for p in csvs:
        print(f"wrote {p.relative_to(ROOT)}")
    for p in split:
        print(f"wrote {p.relative_to(ROOT)}")
